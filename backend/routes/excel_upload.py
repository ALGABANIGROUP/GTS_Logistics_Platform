# backend/routes/excel_upload.py
from typing import Any, List

from fastapi import APIRouter, UploadFile, File, HTTPException
from openpyxl import load_workbook

router = APIRouter()


@router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    # Read uploaded file into a temporary path
    contents = await file.read()
    temp_path = "temp.xlsx"

    with open(temp_path, "wb") as f:
        f.write(contents)

    # Load workbook safely
    try:
        wb = load_workbook(temp_path, data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Excel file: {e}",
        )

    sheet = wb.active
    if sheet is None:
        raise HTTPException(
            status_code=400,
            detail="No active sheet found in the workbook.",
        )

    # Collect rows as plain lists for JSON serialization
    rows: List[List[Any]] = [list(r) for r in sheet.iter_rows(values_only=True)]

    return {
        "message": "Excel processed",
        "rows": rows,
    }
